import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import simpledialog
import subprocess
import dns.resolver
import openpyxl
import OpenSSL.crypto
import requests
import ssl
import whois
import xml.etree.ElementTree as ET


class DataProcessor:
    def __init__(self):
        self.data = None
        self.results = []

    def load_data_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All Files", "*.*")],
        )
        if file_path:
            try:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                self.data = []
                for row in sheet.iter_rows(values_only=True):
                    self.data.append(row[0])
                messagebox.showinfo("Success", "Data loaded from Excel file.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load data from Excel file.\n{str(e)}")
        else:
            messagebox.showwarning("No File Selected", "No file selected.")

    def load_data_from_terminal(self):
        data_input = simpledialog.askstring("Terminal Input", "Enter data (one entry per line, press 'enter' twice to continue):")
        self.data = data_input.splitlines() if data_input else []

    def perform_whois_lookup(self):
        for domain in self.data:
            try:
                w = whois.whois(domain)
                result = f"Domain: {domain}\nRegistrar: {w.registrar}\nName Servers: {', '.join(w.name_servers)}"
            except Exception as e:
                result = f"Whois lookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def perform_nslookup(self):
        for domain in self.data:
            try:
                answers = dns.resolver.resolve(domain, "NS")
                name_servers = [str(rdata) for rdata in answers]
                result = f"Domain: {domain}\nName Servers: {', '.join(name_servers)}"
            except Exception as e:
                result = f"NSLookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def perform_dns_lookup(self):
        for domain in self.data:
            try:
                answers = dns.resolver.resolve(domain)
                records = [str(rdata) for rdata in answers]
                result = f"Domain: {domain}\nDNS Records: {', '.join(records)}"
            except Exception as e:
                result = f"DNS lookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def perform_reverse_dns_lookup(self):
        for ip_address in self.data:
            try:
                answers = dns.resolver.resolve_address(ip_address)
                hostnames = [str(rdata) for rdata in answers]
                result = f"IP Address: {ip_address}\nHostnames: {', '.join(hostnames)}"
            except Exception as e:
                result = f"Reverse DNS lookup failed for IP address {ip_address}: {str(e)}"
            self.results.append(result)

    def perform_ssl_cert_lookup(self):
        for domain in self.data:
            try:
                cert = ssl.get_server_certificate((domain, 443))
                x509 = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, cert)
                expiration_date = x509.get_notAfter().decode()
                result = f"Domain: {domain}\nExpiration Date: {expiration_date}"

                cmd = f"sslscan {domain}"
                output = subprocess.check_output(cmd, shell=True).decode()
                result += f"\n\n{output}"
            except Exception as e:
                result = f"SSL certificate lookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def perform_http_headers_lookup(self):
        for url in self.data:
            try:
                if not url.startswith("http://") and not url.startswith("https://"):
                    url = "http://" + url

                response = requests.head(url)
                headers = response.headers
                result = f"URL: {url}\nHeaders: {headers}"
            except Exception as e:
                result = f"HTTP headers lookup failed for URL {url}: {str(e)}"
            self.results.append(result)

    def perform_subdomain_lookup(self):
        for domain in self.data:
            try:
                cmd = f"amass enum -d {domain}"
                output = subprocess.check_output(cmd, shell=True).decode()
                result = f"Domain: {domain}\nSubdomains:\n\n{output}"
            except Exception as e:
                result = f"Subdomain lookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def save_output_as_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All Files", "*.*")],
        )
        if file_path:
            try:
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.append(["Results"])
                for result in self.results:
                    sheet.append([result])
                wb.save(file_path)
                messagebox.showinfo("Success", "Output saved as Excel file.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save output as Excel file.\n{str(e)}")
        else:
            messagebox.showwarning("No File Selected", "No file selected.")

    def save_output_as_text(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All Files", "*.*")],
        )
        if file_path:
            try:
                with open(file_path, "w") as file:
                    file.write("\n\n".join(self.results))
                messagebox.showinfo("Success", "Output saved as Text file.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save output as Text file.\n{str(e)}")
        else:
            messagebox.showwarning("No File Selected", "No file selected.")

    def save_output_as_xml(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xml",
            filetypes=[("XML files", "*.xml"), ("All Files", "*.*")],
        )
        if file_path:
            try:
                root = ET.Element("Results")
                for result in self.results:
                    element = ET.SubElement(root, "Result")
                    element.text = result
                tree = ET.ElementTree(root)
                tree.write(file_path)
                messagebox.showinfo("Success", "Output saved as XML file.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save output as XML file.\n{str(e)}")
        else:
            messagebox.showwarning("No File Selected", "No file selected.")

    def print_output_to_terminal(self):
        output_text = "Output:\n\n" + "\n\n".join(self.results)
        messagebox.showinfo("Output", output_text)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Domain Tool")
        self.geometry("400x500")

        self.data_processor = DataProcessor()

        btn_load_excel = tk.Button(self, text="Load Excel File", command=self.load_excel_file)
        btn_load_excel.pack(pady=10)

        btn_terminal_input = tk.Button(self, text="Terminal Input", command=self.load_terminal_input)
        btn_terminal_input.pack(pady=10)

        btn_whois = tk.Button(self, text="Perform WHOIS Lookup", command=self.perform_whois_lookup)
        btn_whois.pack(pady=10)

        btn_nslookup = tk.Button(self, text="Perform NSLookup", command=self.perform_nslookup)
        btn_nslookup.pack(pady=10)

        btn_dns_lookup = tk.Button(self, text="Perform DNS Lookup", command=self.perform_dns_lookup)
        btn_dns_lookup.pack(pady=10)

        btn_reverse_dns_lookup = tk.Button(self, text="Perform Reverse DNS Lookup", command=self.perform_reverse_dns_lookup)
        btn_reverse_dns_lookup.pack(pady=10)

        btn_ssl_cert_lookup = tk.Button(self, text="Perform SSL Certificate Lookup", command=self.perform_ssl_cert_lookup)
        btn_ssl_cert_lookup.pack(pady=10)

        btn_http_headers_lookup = tk.Button(self, text="Perform HTTP Headers Lookup", command=self.perform_http_headers_lookup)
        btn_http_headers_lookup.pack(pady=10)

        btn_subdomain_lookup = tk.Button(self, text="Perform Subdomain Lookup", command=self.perform_subdomain_lookup)
        btn_subdomain_lookup.pack(pady=10)

        btn_save_excel = tk.Button(self, text="Save Output as Excel", command=self.save_output_as_excel)
        btn_save_excel.pack(pady=10)

        btn_save_text = tk.Button(self, text="Save Output as Text", command=self.save_output_as_text)
        btn_save_text.pack(pady=10)

        btn_save_xml = tk.Button(self, text="Save Output as XML", command=self.save_output_as_xml)
        btn_save_xml.pack(pady=10)

        btn_print_output = tk.Button(self, text="Print Output", command=self.print_output)
        btn_print_output.pack(pady=10)

        btn_exit = tk.Button(self, text="Exit", command=self.destroy)
        btn_exit.pack(pady=10)

    def load_excel_file(self):
        self.data_processor.load_data_from_excel()

    def load_terminal_input(self):
        self.data_processor.load_data_from_terminal()

    def perform_whois_lookup(self):
        self.data_processor.perform_whois_lookup()

    def perform_nslookup(self):
        self.data_processor.perform_nslookup()

    def perform_dns_lookup(self):
        self.data_processor.perform_dns_lookup()

    def perform_reverse_dns_lookup(self):
        self.data_processor.perform_reverse_dns_lookup()

    def perform_ssl_cert_lookup(self):
        self.data_processor.perform_ssl_cert_lookup()

    def perform_http_headers_lookup(self):
        self.data_processor.perform_http_headers_lookup()

    def perform_subdomain_lookup(self):
        self.data_processor.perform_subdomain_lookup()

    def save_output_as_excel(self):
        self.data_processor.save_output_as_excel()

    def save_output_as_text(self):
        self.data_processor.save_output_as_text()

    def save_output_as_xml(self):
        self.data_processor.save_output_as_xml()

    def print_output(self):
        self.data_processor.print_output_to_terminal()


if __name__ == "__main__":
    app = App()
    app.mainloop()
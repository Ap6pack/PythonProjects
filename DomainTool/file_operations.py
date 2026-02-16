import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import xml.etree.ElementTree as ET
from tkinter import filedialog
import tkinter as tk
from network_operations import format_results_for_text

def load_data_from_excel(file_path):
    if not file_path:
        print("No file selected. Exiting...")
        return False
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row[0])
        print(f"Data loaded from Excel file: {file_path}")
        return data
    except Exception as e:
        raise ValueError(f"Failed to load data from Excel file. Error: {str(e)}")

def load_data_from_xml(file_path):
    if not file_path:
        print("No file selected. Exiting...")
        return False
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        data = [elem.text for elem in root.findall(".//domain")]
        print(f"Data loaded from XML file: {file_path}")
        return data
    except Exception as e:
        raise ValueError("Failed to load data from XML file. Error: " + str(e))

def load_data_from_text(file_path):
    if not file_path:
        print("No file selected. Exiting...")
        return False
    try:
        with open(file_path, "r") as file:
            data = file.read().splitlines()
        print(f"Data loaded from Text file: {file_path}")
        return data
    except Exception as e:
        print(f"Failed to load data from Text file. Error: {str(e)}")
        return

def load_data_from_terminal(data_processor):
    print("Enter data (one entry per line, press 'enter' twice to continue):")
    data_processor.clear_data()  # Clear the data before loading new data
    entry = input()

    while entry or not data_processor.data:
        if not entry.strip():
            print("Domain not entered. Please try again or enter 'b' to go back to the main menu.")
            entry = input()
            if entry.strip() == 'b':
                return False  # Return False to indicate going back to the main menu
        else:
            data_processor.data.append(entry)
            entry = input()
    return True  # Return True to indicate successful data loading


def save_output_as_excel(data_processor):
    """
    Save results to Excel with structured columns.
    Automatically detects the lookup type and formats accordingly.
    """
    try:
        results = data_processor.results
        if not results:
            print("No results to save.")
            return
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # Detect lookup type based on keys in first result
        if isinstance(results[0], dict):
            # Structured data - create columns
            headers = list(results[0].keys())
            
            # Apply header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Write headers
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header.replace('_', ' ').title())
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data rows
            for row_num, result in enumerate(results, start=2):
                for col_num, header in enumerate(headers, start=1):
                    value = result[header]
                    
                    # Handle nested dictionaries (like HTTP headers)
                    if isinstance(value, dict):
                        value = '\n'.join([f"{k}: {v}" for k, v in value.items()])
                    
                    # Sanitize value for Excel
                    if value is not None:
                        value_str = str(value)
                        # Remove illegal characters
                        value_str = value_str.replace("\r", "").replace("\x00", "")
                        sheet.cell(row=row_num, column=col_num, value=value_str)
                    else:
                        sheet.cell(row=row_num, column=col_num, value="")
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
                sheet.column_dimensions[column_letter].width = adjusted_width
            
        else:
            # Legacy format - single column
            for i, result in enumerate(results, start=1):
                result = str(result).replace("\r", "").replace("\n", " ")
                try:
                    sheet.cell(row=i, column=1, value=result)
                except openpyxl.utils.exceptions.IllegalCharacterError:
                    sanitized_result = "".join(c for c in result if c.isprintable())
                    sheet.cell(row=i, column=1, value=sanitized_result)
        
        # Get file path and save
        file_path = get_file_path("xlsx", "save")
        if not file_path:
            print("File selection canceled. Returning to Main Menu.")
            return
        
        wb.save(file_path)
        wb.close()
        print(f"✓ Output saved as Excel file: {file_path}")
        
    except Exception as e:
        print(f"Failed to save output as Excel file. Error: {str(e)}")


def save_output_as_xml(data_processor):
    """
    Save results to XML format.
    """
    try:
        results = data_processor.results
        root = ET.Element("results")
        
        for result in results:
            if isinstance(result, dict):
                # Structured data
                result_elem = ET.SubElement(root, "result")
                for key, value in result.items():
                    elem = ET.SubElement(result_elem, key)
                    if isinstance(value, dict):
                        # Handle nested dictionaries
                        for sub_key, sub_value in value.items():
                            sub_elem = ET.SubElement(elem, sub_key)
                            sub_elem.text = str(sub_value) if sub_value is not None else ""
                    else:
                        elem.text = str(value) if value is not None else ""
            else:
                # Legacy string format
                result_elem = ET.SubElement(root, "result")
                result_elem.text = str(result)
        
        tree = ET.ElementTree(root)
        file_path = get_file_path("xml", "save")
        if not file_path:
            print("File selection canceled. Returning to Main Menu.")
            return
        
        tree.write(file_path, encoding="UTF-8", xml_declaration=True)
        print(f"✓ Output saved as XML file: {file_path}")
        
    except Exception as e:
        print(f"Failed to save output as XML file. Error: {str(e)}")


def save_output_as_text(data_processor):
    """
    Save results to text file with readable formatting.
    """
    try:
        results = data_processor.results
        
        file_path = get_file_path("txt", "save")
        if not file_path:
            print("File selection canceled. Returning to Main Menu.")
            return
        
        with open(file_path, "w") as file:
            for result in results:
                if isinstance(result, dict):
                    # Format structured data as text
                    lookup_type = data_processor.current_lookup_type if hasattr(data_processor, 'current_lookup_type') else 'whois'
                    formatted = format_results_for_text([result], lookup_type)
                    file.write(formatted[0] + "\n\n")
                else:
                    # Legacy string format
                    file.write(str(result) + "\n\n")
        
        print(f"✓ Output saved as Text file: {file_path}")
        
    except Exception as e:
        print(f"Failed to save output as Text file. Error: {str(e)}")


def get_file_path(file_type, action):
    root = tk.Tk()
    root.withdraw()

    if action == "load":
        file_path = filedialog.askopenfilename(filetypes=[(f"{file_type} files", f"*.{file_type.lower()}")])
        if not file_path:
            print("File selection canceled. Returning to Main Menu.")
            return None
    elif action == "save":
        file_path = filedialog.asksaveasfilename(
            defaultextension=f".{file_type.lower()}",
            filetypes=[(f"{file_type} files", f"*.{file_type.lower()}")]
        )
        if not file_path:
            print("No file selected. Returning to Main Menu.")
            return None
    else:
        print("Invalid action. Supported actions are 'load' and 'save'.")

    return file_path


def print_output_to_terminal(data_processor):
    """
    Print results to terminal with readable formatting.
    """
    print("\n" + "="*60)
    print("OUTPUT RESULTS")
    print("="*60 + "\n")
    
    results = data_processor.results
    
    for i, result in enumerate(results, start=1):
        if isinstance(result, dict):
            # Format structured data
            lookup_type = data_processor.current_lookup_type if hasattr(data_processor, 'current_lookup_type') else 'whois'
            formatted = format_results_for_text([result], lookup_type)
            print(f"[{i}] {formatted[0]}")
        else:
            # Legacy string format
            print(f"[{i}] {result}")
        
        print("-" * 60)
    
    print()

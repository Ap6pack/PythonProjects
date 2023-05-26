import sys
import subprocess
import tkinter as tk
from tkinter import Tk
from tkinter import filedialog
from tkinter.filedialog import askopenfilename, asksaveasfilename

import dns.resolver
import openpyxl
import OpenSSL.crypto
import requests
import ssl
import whois
import xml.etree.ElementTree as ET


class DataProcessor:
    def __init__(self):
        self.data = None
        self.results = []

    def load_data_from_excel(self, file_path):
        if not file_path:
            print("No file selected. Exiting...")
            return False
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            self.data = []
            for row in sheet.iter_rows(values_only=True):
                self.data.append(row[0])
            print(f"Data loaded from Excel file: {file_path}")
            return True
        except Exception as e:
            raise ValueError(f"Failed to load data from Excel file. Error: {str(e)}")
   
    def load_data_from_xml(self, file_path):
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            self.data = [elem.text for elem in root.findall(".//data")]
        except Exception as e:
            raise ValueError("Failed to load data from XML file. Error: " + str(e))

    def load_data_from_text(self, file_path):
        try:
            with open(file_path, "r") as file:
                self.data = file.read().splitlines()
        except Exception as e:
            raise ValueError("Failed to load data from Text file. Error: " + str(e))

    def load_data_from_terminal(self):
        print("Enter data (one entry per line, press 'enter' twice to continue):")
        self.data = []
        entry = input()
        while entry:
            self.data.append(entry)
            entry = input()

    def perform_functions(self, function_options):
        self.results = []  # Clear the results list
        function_options = function_options.strip()
        for option in function_options:
            option = option.strip()
            if option == "a":
                self.perform_whois_lookup()
            elif option == "b":
                self.perform_nslookup()
            elif option == "c":
                self.perform_dns_lookup()
            elif option == "d":
                self.perform_reverse_dns_lookup()
            elif option == "e":
                self.perform_ssl_cert_lookup()
            elif option == "f":
                self.perform_http_headers_lookup()
            elif option == "g":
                self.perform_subdomain_lookup()
            elif option == "h":
                self._show_menu()

    def perform_whois_lookup(self):
        for domain in self.data:
            try:
                w = whois.whois(domain)
                result = f"Domain: {domain}\nRegistrar: {w.registrar}\nName Servers: {', '.join(w.name_servers)}"
            except Exception as e:
                result = f"Whois lookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def perform_nslookup(self):
        for domain in self.data:
            try:
                answers = dns.resolver.resolve(domain, "NS")
                name_servers = [str(rdata) for rdata in answers]
                result = f"Domain: {domain}\nName Servers: {', '.join(name_servers)}"
            except Exception as e:
                result = f"NSLookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def perform_dns_lookup(self):
        for domain in self.data:
            try:
                answers = dns.resolver.resolve(domain)
                records = [str(rdata) for rdata in answers]
                result = f"Domain: {domain}\nDNS Records: {', '.join(records)}"
            except Exception as e:
                result = f"DNS lookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def perform_reverse_dns_lookup(self):
        for ip_address in self.data:
            try:
                answers = dns.resolver.resolve_address(ip_address)
                hostnames = [str(rdata) for rdata in answers]
                result = f"IP Address: {ip_address}\nHostnames: {', '.join(hostnames)}"
            except Exception as e:
                result = (
                    f"Reverse DNS lookup failed for IP address {ip_address}: {str(e)}"
                )
            self.results.append(result)

    def perform_ssl_cert_lookup(self):
        for domain in self.data:
            try:
                cert = ssl.get_server_certificate((domain, 443))
                x509 = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, cert)
                expiration_date = x509.get_notAfter().decode()
                result = f"Domain: {domain}\nExpiration Date: {expiration_date}"

                cmd = f"sslscan {domain}"
                sslscan_output = subprocess.check_output(cmd, shell=True, universal_newlines=True)
                result += f"\n\nSSL Scan:\n{sslscan_output}"
            except Exception as e:
                result = f"SSL certificate lookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def perform_http_headers_lookup(self):
        for url in self.data:
            try:
                if not url.startswith("http://") and not url.startswith("https://"):
                    url = "http://" + url

                response = requests.head(url)
                headers = response.headers
                result = f"URL: {url}\nHeaders:"
                
                for header, value in headers.items():
                    result += f"\n\t{header}: {value}"

                self.results.append(result)
            except Exception as e:
                result = f"HTTP headers lookup failed for URL {url}: {str(e)}"
                self.results.append(result)

    def perform_subdomain_lookup(self):
        for domain in self.data:
            try:
                output = subprocess.check_output(["dig", f"{domain}", "+short", "+sub"])
                subdomains = output.decode().strip().splitlines()
                result = f"Domain: {domain}\nSubdomains: {', '.join(subdomains)}"
            except Exception as e:
                result = f"Subdomain lookup failed for domain {domain}: {str(e)}"
            self.results.append(result)

    def save_output_as_excel(self):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active

            for i, result in enumerate(self.results, start=1):
                # Replace illegal characters with a placeholder
                result = result.replace('\r', '').replace('\n', ' ')
                try:
                    sheet.cell(row=i, column=1, value=result)
                except openpyxl.utils.exceptions.IllegalCharacterError:
                    # Handle the error by replacing illegal characters with an empty string
                    sanitized_result = ''.join(c for c in result if c.isprintable())
                    sheet.cell(row=i, column=1, value=sanitized_result)

            file_path = self._get_save_file_path("xlsx")
            wb.save(file_path)
            wb.close()
            print(f"Output saved as Excel file: {file_path}")
        except Exception as e:
            raise ValueError("Failed to save output as Excel file. Error: " + str(e))
    
    def save_output_as_text(self):
        try:
            file_path = self._get_save_file_path("txt")
            with open(file_path, "w") as file:
                for result in self.results:
                    file.write(result + "\n\n")
            print(f"Output saved as Text file: {file_path}")
        except Exception as e:
            raise ValueError("Failed to save output as Text file. Error: " + str(e))

    def save_output_as_xml(self):
        try:
            root = ET.Element("results")
            for result in self.results:
                elem = ET.SubElement(root, "result")
                elem.text = result
            tree = ET.ElementTree(root)
            file_path = self._get_save_file_path("xml")
            tree.write(file_path)
            print(f"Output saved as XML file: {file_path}")
        except Exception as e:
            raise ValueError("Failed to save output as XML file. Error: " + str(e))

    def _get_save_file_path(self, file_type):
        root = Tk()
        root.withdraw()
        file_path = filedialog.asksaveasfilename(
            defaultextension=f".{file_type.lower()}",
            filetypes=[(f"{file_type} files", f"*.{file_type.lower()}")],
        )
        return file_path
    
    def print_output_to_terminal(self):
        print("Output:")
        for item in self.results:
            print(item)
            
def main():
    dp = DataProcessor()
    should_exit = False

    while not should_exit:
        print("\nSelect data source:")
        print("1. Excel file")
        print("2. XML file")
        print("3. Text file")
        print("4. Terminal input")
        print("5. Exit")
        source_option = input("Enter option number: ")

        if source_option == "1":
            Tk().withdraw()  # Hide the root window
            file_path = askopenfilename(
                title="Select Excel file",
                filetypes=[("Excel files", "*.xlsx"), ("All Files", "*.*")],
            )
            if file_path:
                if dp.load_data_from_excel(file_path):
                    dp.file_type = "Excel"
            else:
                print("No file selected.")
                continue
        elif source_option == "2":
            Tk().withdraw()
            file_path = askopenfilename(title="Select XML file")
            dp.load_data_from_xml(file_path)
        elif source_option == "3":
            Tk().withdraw()
            file_path = askopenfilename(title="Select text file")
            dp.load_data_from_text(file_path)
        elif source_option == "4":
            dp.load_data_from_terminal()
        elif source_option == "5":
            should_exit = True
        else:
            print("Invalid option. Please choose a valid option.")
            continue

        if not should_exit:
            print("\nSelect functions to perform:")
            print("a. WHOIS lookup")
            print("b. NSLookup")
            print("c. DNS lookup")
            print("d. Reverse DNS lookup")
            print("e. SSL certificate lookup")
            print("f. HTTP headers lookup")
            print("g. Subdomain lookup")
            print("h. Show menu")
            print("i. Save output as Excel file")
            print("j. Save output as Text file")
            print("k. Save output as XML file")
            print("l. Print output to terminal")
            print("m. Exit")

            function_options = input("Enter options (e.g., 'abce'): ")

            if "h" in function_options:
                dp._show_menu()
            elif "i" in function_options:
                dp.save_output_as_excel()
            elif "j" in function_options:
                dp.save_output_as_text()
            elif "k" in function_options:
                dp.save_output_as_xml()
            elif "l" in function_options:
                dp.print_output_to_terminal()
            elif "m" in function_options:
                should_exit = True
            else:
                dp.perform_functions(function_options)

if __name__ == "__main__":
    main()
